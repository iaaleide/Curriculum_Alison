#!/usr/bin/env python3
"""Merge researched emails and generate organized Excel workbook."""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path("/workspace/alb_emails_research")
OUT_XLSX = Path("/workspace/ALB_Brasil_Emails_Empresas.xlsx")
OUT_DIR = Path("/opt/cursor/artifacts")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def norm(s: str) -> str:
    s = s or ""
    # Source CSV contains U+FFFD where çã/ã/é were lost
    s = s.replace("Nutri\ufffd\ufffdo", "Nutricao").replace("nutri\ufffd\ufffdo", "nutricao")
    s = s.replace("Milh\ufffdo", "Milhao").replace("milh\ufffdo", "milhao")
    s = s.replace("Pecu\ufffdria", "Pecuaria").replace("Ra\ufffd\ufffdes", "Racoes")
    s = s.replace("\ufffd", "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    s = s.replace("ç", "c").replace("ñ", "n")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = s.replace("nutri o ", "nutricao ").replace("pecu ria", "pecuaria")
    # collapsed forms after lost ã
    s = re.sub(r"\bmilho ingredients\b", "milhao ingredients", s)
    return re.sub(r"\s+", " ", s).strip()


def load_curated() -> dict:
    curated = {}
    for name in ("curated_emails.json", "extra_curated.json", "batch2_curated.json"):
        p = ROOT / name
        if p.exists():
            data = json.loads(p.read_text(encoding="utf-8"))
            for k, v in data.items():
                curated[norm(k)] = v
    return curated


def _stem_tokens(tokens: list[str]) -> list[str]:
    """Loose stems to tolerate lost accents from source CSV (comrcio≈comercio)."""
    stop = {"de", "da", "do", "das", "dos", "e", "ltda", "ltd", "sa", "s", "a", "o", "i"}
    out = []
    for t in tokens:
        if t in stop or len(t) < 2:
            continue
        # keep first 6 chars as stem for longer words
        out.append(t[:6] if len(t) > 6 else t)
    return out


def find_curated(empresa: str, curated: dict) -> dict | None:
    key = norm(empresa)
    if key in curated:
        return curated[key]
    tokens = key.split()
    stems = _stem_tokens(tokens)
    SHORT = {"brf", "jbs", "ldc", "bunge", "cargill", "comigo", "guabi", "inpasa", "marfrig", "minerva"}
    candidates = []
    for ck, cv in curated.items():
        ct = ck.split()
        if not ct:
            continue
        if key == ck:
            candidates.append((100, cv, ck))
            continue
        # multi-word curated key must match company prefix tokens
        if len(ct) >= 2 and tokens[: len(ct)] == ct:
            candidates.append((95, cv, ck))
            continue
        # stem-based prefix match (handles Milhao/Milho, Guzzo Agronegocios, etc.)
        cstems = _stem_tokens(ct)
        if len(cstems) >= 2 and len(stems) >= len(cstems) and stems[: len(cstems)] == cstems:
            candidates.append((90, cv, ck))
            continue
        # first two significant stems equal
        if len(cstems) >= 2 and len(stems) >= 2 and stems[:2] == cstems[:2] and len(cstems[0]) >= 4:
            candidates.append((88, cv, ck))
            continue
        # single-token curated key
        if len(ct) == 1:
            tok = ct[0]
            if tok in SHORT:
                if tokens and tokens[0] == tok:
                    if tok == "cargill" and len(tokens) > 1 and tokens[1] == "bioenergia":
                        continue
                    candidates.append((85, cv, ck))
            elif tok == tokens[0] and len(tok) >= 5:
                candidates.append((70, cv, ck))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def extract_phones(raw: str) -> list[str]:
    """Extract BR phone numbers as digits-only (DDD+number)."""
    if not raw:
        return []
    text = raw.replace("Whtas app", " ").replace("Whats app", " ").replace("WhatsApp", " ")
    text = re.sub(r"(?i)whtas?\s*app", " ", text)
    # normalize separators
    chunks = re.split(r"[|/]|;|,|\bn/\b|\be\b", text)
    found = []
    for ch in chunks:
        digits = re.sub(r"\D", "", ch)
        if digits.startswith("55") and len(digits) >= 12:
            digits = digits[2:]
        # 0800
        if digits.startswith("0800") and len(digits) >= 10:
            found.append(digits[:11] if len(digits) >= 11 else digits)
            continue
        # DDD + 8 or 9
        if len(digits) >= 10:
            # take last 10 or 11
            if len(digits) >= 11 and digits[-9] == "9":
                found.append(digits[-11:])
            else:
                found.append(digits[-10:])
        elif len(digits) == 9 and digits.startswith("9"):
            # mobile without DDD - keep as is (incomplete)
            found.append(digits)
    # unique preserve order
    out = []
    for n in found:
        if n not in out:
            out.append(n)
    return out


def format_phone_br(digits: str) -> str:
    if not digits:
        return ""
    if digits.startswith("0800") and len(digits) >= 10:
        d = digits.ljust(11, "0")[:11]
        return f"0800 {d[4:7]} {d[7:]}"
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return digits


def pick_whatsapp(phones: list[str]) -> tuple[str, str]:
    """Return (whatsapp formatted, wa.me link) preferring mobile 9-digit."""
    mobiles = [p for p in phones if len(p) == 11 and p[2] == "9"]
    # also 10-digit landline less preferred
    chosen = mobiles[0] if mobiles else (phones[0] if phones and not phones[0].startswith("0800") else "")
    if not chosen:
        # fallback first mobile-like from any
        for p in phones:
            if "9" in p[2:3] or (len(p) == 11):
                chosen = p
                break
    if not chosen:
        return "", ""
    link = f"https://wa.me/55{chosen}" if not chosen.startswith("0800") else ""
    return format_phone_br(chosen), link


def is_bad_auto(email: str, empresa: str) -> bool:
    if not email:
        return True
    e = email.lower()
    # known wrong mappings from CSV website mixups
    bad_pairs = [
        ("vibra.com", "jbs"),
        ("vibra.com", "seara"),
        ("riobrancoalimentossa.com", "pif paf"),
        ("gmail.com", "special dog"),
        ("jatainet.com.br", ""),
    ]
    en = norm(empresa)
    for domain, must in bad_pairs:
        if domain in e:
            if not must or must in en:
                return True
    return False


def fix_display_name(s: str) -> str:
    """Repair mojibake/replacement chars from source CSV for Excel display."""
    if not s:
        return s
    # U+FFFD pairs commonly standing for çã / ã / é etc in this file
    replacements = [
        ("Nutri\ufffd\ufffdo", "Nutrição"),
        ("nutri\ufffd\ufffdo", "nutrição"),
        ("Milh\ufffdo", "Milhão"),
        ("milh\ufffdo", "milhão"),
        ("Ra\ufffd\ufffdes", "Rações"),
        ("ra\ufffd\ufffdes", "rações"),
        ("Com\ufffdrcio", "Comércio"),
        ("Exporta\ufffd\ufffdo", "Exportação"),
        ("Agroneg\ufffdcios", "Agronegócios"),
        ("Agrolog\ufffdstica", "Agrologística"),
        ("S\ufffdo", "São"),
        ("s\ufffdo", "são"),
        ("S\ufffd", "Sá"),
        ("Goi\ufffds", "Goiás"),
        ("goi\ufffds", "goiás"),
        ("Paran\ufffd", "Paraná"),
        ("Esp\ufffdrito", "Espírito"),
        ("Uberl\ufffdndia", "Uberlândia"),
        ("Itamb\ufffd", "Itambé"),
        ("Pecu\ufffdria", "Pecuária"),
        ("Ind\ufffdstria", "Indústria"),
        ("ind\ufffdstria", "indústria"),
        ("Armaz\ufffdm", "Armazém"),
        ("armaz\ufffdm", "armazém"),
        ("Armaz\ufffdns", "Armazéns"),
        ("F\ufffdbrica", "Fábrica"),
        ("f\ufffdbrica", "fábrica"),
        ("Jos\ufffd", "José"),
        ("C\ufffdu", "Céu"),
        ("Chapad\ufffdo", "Chapadão"),
        ("An\ufffdpolis", "Anápolis"),
        ("Goian\ufffdsia", "Goianésia"),
        ("Ot\ufffdvio", "Otávio"),
        ("Abadi\ufffdnia", "Abadiânia"),
        ("Vicentin\ufffdpolis", "Vicentinópolis"),
        ("Castel\ufffdndia", "Castelândia"),
        ("Mauril\ufffdndia", "Maurilândia"),
        ("Auril\ufffdndia", "Aurilândia"),
        ("Naz\ufffdrio", "Nazário"),
        ("Goian\ufffdpolis", "Goianápolis"),
        ("Luzi\ufffdnia", "Luziânia"),
        ("Jata\ufffd", "Jataí"),
        ("Para\ufffdna", "Paraúna"),
        ("Itabera\ufffd", "Itaberaí"),
        ("Quir\ufffdnopolis", "Quirinópolis"),
        ("Montividiu", "Montividiu"),
        ("Sert\ufffdozinho", "Sertãozinho"),
        ("Guai\ufffdara", "Guaiçara"),
        ("Marin\ufffdpolis", "Marinópolis"),
        ("Guzol\ufffdndia", "Guzolândia"),
        ("Aprez\ufffdvel", "Aprazível"),
        ("Rondonopolis", "Rondonópolis"),
        ("Rondon\ufffdpolis", "Rondonópolis"),
        ("Irine\ufffdpolis", "Irineópolis"),
        ("Jacito", "Jacinto"),
        ("Chopizinho", "Chopinzinho"),
        ("Vit\ufffdria", "Vitória"),
        ("Gr\ufffdo", "Grão"),
        ("Uni\ufffdo", "União"),
        ("Alian\ufffda", "Aliança"),
        ("Agrolog\ufffdstica", "Agrologística"),
        ("Av\ufffdcola", "Avícola"),
        ("av\ufffdcola", "avícola"),
        ("M\ufffde", "Mãe"),
        ("Salom\ufffdo", "Salomão"),
        ("Inac\ufffdo", "Inácio"),
        ("Jetib\ufffd", "Jetibá"),
        ("Saager", "Saager"),
        ("Nakanishi", "Nakanishi"),
        ("Altin\ufffdpolis", "Altinópolis"),
        ("Ip\ufffd", "Ipê"),
        ("Jaguar\ufffd", "Jaguari"),
        ("Suzano", "Suzano"),
        ("Tatu\ufffd", "Tatuí"),
        ("Avar\ufffd", "Avaré"),
        ("Tiete", "Tietê"),
        ("Gr\ufffdo Par\ufffd", "Grão Pará"),
        ("Par\ufffd", "Pará"),
        ("f\ufffd", "fé"),
        ("F\ufffd", "Fé"),
        ("Confnamento", "Confinamento"),
        ("Boithtel", "Boitel"),
        ("Boihtel", "Boitel"),
        ("\ufffd", ""),  # drop leftover replacement chars
    ]
    out = s
    for a, b in replacements:
        out = out.replace(a, b)
    return out


def merge():
    empresas = json.loads((ROOT / "empresas.json").read_text(encoding="utf-8"))
    auto_list = []
    if (ROOT / "emails_encontrados.json").exists():
        auto_list = json.loads((ROOT / "emails_encontrados.json").read_text(encoding="utf-8"))
    auto = {x["id"]: x for x in auto_list}
    curated = load_curated()

    rows = []
    for c in empresas:
        a = auto.get(c["id"], {})
        cur = find_curated(c["empresa"], curated) or {}

        email = ""
        emails = []
        fonte = ""
        conf = "-"
        obs = c.get("obs") or ""
        website = c.get("website") or ""

        if cur:
            email = cur.get("email_principal") or ""
            emails = cur.get("emails_encontrados") or ([] if not email else [email])
            fonte = cur.get("fonte") or "Curadoria"
            conf = cur.get("confianca") or ("Alta" if email else "-")
            if cur.get("website_corrigido"):
                website = cur["website_corrigido"]
            if cur.get("obs_contato"):
                obs = cur["obs_contato"]
        else:
            email = a.get("email_principal") or ""
            emails = a.get("emails_encontrados") or []
            fonte = a.get("fonte") or ""
            conf = a.get("confianca") or "-"
            if is_bad_auto(email, c["empresa"]):
                email = ""
                emails = [e for e in emails if not is_bad_auto(e, c["empresa"])]
                email = emails[0] if emails else ""
                if not email:
                    conf = "-"
                    fonte = ""

        tel_raw = (c.get("telefone") or "").strip()
        phones = extract_phones(tel_raw)
        tel_fmt = " / ".join(format_phone_br(p) for p in phones) if phones else tel_raw
        wa_num, wa_link = pick_whatsapp(phones)
        # if source mentions whatsapp explicitly and we have a mobile, mark it
        mentions_wa = bool(re.search(r"(?i)whats|whtas", tel_raw)) or bool(wa_num)

        if email:
            status = "E-mail encontrado"
            canal = "E-mail"
        elif wa_num:
            status = "Contato por Telefone/WhatsApp"
            canal = "WhatsApp/Telefone"
        elif tel_raw:
            status = "Contato por Telefone"
            canal = "Telefone"
        elif website:
            status = "Somente website/formulário"
            canal = "Website"
        else:
            status = "Sem contato público"
            canal = "-"

        if email and conf == "-":
            conf = "Média"

        rows.append(
            {
                "id": c["id"],
                "empresa": fix_display_name(c["empresa"]),
                "uf": fix_display_name(c.get("uf") or ""),
                "municipio": fix_display_name(c.get("municipio") or ""),
                "segmento": fix_display_name(c.get("segmento") or ""),
                "area": fix_display_name(c.get("area") or ""),
                "perfil": fix_display_name(c.get("perfil") or ""),
                "email_principal": email,
                "emails_adicionais": "; ".join([e for e in emails if e != email]),
                "telefone": tel_fmt,
                "telefone_original": tel_raw,
                "whatsapp": wa_num,
                "whatsapp_link": wa_link if mentions_wa or wa_num else "",
                "website": website,
                "canal_preferencial": canal,
                "status": status,
                "confianca": conf if email else "-",
                "fonte": fonte,
                "obs_contato": obs,
                "ocorrencias_planilha": c.get("ocorrencias", 1),
            }
        )
    return rows


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, len(val))
        ws.column_dimensions[letter].width = min(max(12, length + 2), max_width)


def write_sheet(ws, headers, data_rows, key_order):
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(data_rows, 2):
        for c_i, key in enumerate(key_order, 1):
            val = row.get(key, "")
            cell = ws.cell(r_i, c_i, val)
            if key == "whatsapp_link" and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"
    style_header(ws, len(headers))
    autosize(ws)
    # zebra for data
    zebra = PatternFill("solid", fgColor="F2F2F2")
    for r_i in range(2, len(data_rows) + 2):
        if r_i % 2 == 0:
            for c_i in range(1, len(headers) + 1):
                cell = ws.cell(r_i, c_i)
                if not cell.hyperlink:
                    cell.fill = zebra
        for c_i in range(1, len(headers) + 1):
            ws.cell(r_i, c_i).alignment = Alignment(vertical="center", wrap_text=True)


def main():
    rows = merge()
    rows_sorted = sorted(rows, key=lambda r: (r["uf"], r["segmento"], r["empresa"]))

    with_email = [r for r in rows_sorted if r["email_principal"]]
    without = [r for r in rows_sorted if not r["email_principal"]]

    wb = Workbook()

    # Resumo
    ws0 = wb.active
    ws0.title = "01_Resumo"
    ws0["A1"] = "ALB Brasil — Contatos das Empresas (E-mail / Telefone / WhatsApp)"
    ws0["A1"].font = Font(bold=True, size=16, color="1F4E79", name="Calibri")
    ws0["A2"] = (
        "Planilha com e-mails públicos pesquisados + telefone/WhatsApp da lista original. "
        "Quando não há e-mail público, o canal principal é telefone ou WhatsApp."
    )
    ws0["A2"].alignment = Alignment(wrap_text=True)
    ws0.merge_cells("A2:F2")
    ws0.row_dimensions[2].height = 40

    com_wa = sum(1 for r in rows if r.get("whatsapp"))
    com_tel = sum(1 for r in rows if r.get("telefone"))
    stats = [
        ("Total de empresas únicas", len(rows)),
        ("Com e-mail encontrado", len(with_email)),
        ("Com telefone", com_tel),
        ("Com WhatsApp (celular identificado)", com_wa),
        ("Sem e-mail — usar Telefone/WhatsApp", sum(1 for r in without if r["telefone"] or r.get("whatsapp"))),
        ("Sem contato público", sum(1 for r in without if not r["telefone"] and not r["website"])),
        ("Confiança Alta (e-mail)", sum(1 for r in with_email if r["confianca"] == "Alta")),
    ]
    ws0["A4"] = "Indicador"
    ws0["B4"] = "Valor"
    for i, (k, v) in enumerate(stats, 5):
        ws0[f"A{i}"] = k
        ws0[f"B{i}"] = v
    style_header(ws0, 2)
    # rewrite header only on row 4
    for col in range(1, 3):
        cell = ws0.cell(4, col)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)

    ws0["A13"] = "Abas"
    ws0["A13"].font = Font(bold=True, size=12, color="1F4E79")
    guia = [
        "02_Contatos_Completos — Empresa | E-mail | Telefone | WhatsApp (use esta)",
        "03_Todos_Detalhado — lista completa com todas as colunas",
        "04_Com_Email — só quem tem e-mail",
        "05_Telefone_WhatsApp — fila telefone/WhatsApp",
        "06_Sem_Email — sem e-mail (usar telefone/WhatsApp)",
        "07_Por_UF — visão por estado",
        "08_Nutricao_Frigorificos — foco ração/frigorífico",
        "09_Lista_Disparo_Email — disparo por e-mail",
        "10_Legenda — critérios de confiança",
    ]
    for i, g in enumerate(guia, 14):
        ws0[f"A{i}"] = g
    ws0["A23"] = "Observações importantes"
    ws0["A23"].font = Font(bold=True, color="C00000")
    ws0["A24"] = (
        "1) Aba principal: 02_Contatos_Completos (E-mail | Telefone | WhatsApp). "
        "2) Sem e-mail: use Telefone/WhatsApp (aba 05). "
        "3) Link WhatsApp abre wa.me. 0800 não vai para WhatsApp. "
        "4) Validar contatos antes de disparos (LGPD)."
    )
    ws0["A24"].alignment = Alignment(wrap_text=True)
    ws0.merge_cells("A24:F24")
    ws0.row_dimensions[24].height = 55
    ws0.column_dimensions["A"].width = 60
    ws0.column_dimensions["B"].width = 18

    headers = [
        "Empresa",
        "UF",
        "Município",
        "Segmento",
        "Área de Atuação",
        "Canal Preferencial",
        "E-mail Principal",
        "E-mails Adicionais",
        "Telefone",
        "WhatsApp",
        "Link WhatsApp",
        "Website",
        "Status",
        "Confiança",
        "Fonte",
        "Observações",
    ]
    keys = [
        "empresa",
        "uf",
        "municipio",
        "segmento",
        "area",
        "canal_preferencial",
        "email_principal",
        "emails_adicionais",
        "telefone",
        "whatsapp",
        "whatsapp_link",
        "website",
        "status",
        "confianca",
        "fonte",
        "obs_contato",
    ]

    # Simple front sheet with the core contact columns
    ws_simple = wb.create_sheet("02_Contatos_Completos")
    simple_headers = [
        "Empresa",
        "UF",
        "Município",
        "Segmento",
        "E-mail",
        "Telefone",
        "WhatsApp",
        "Link WhatsApp",
        "Canal Preferencial",
        "Observações",
    ]
    simple_keys = [
        "empresa",
        "uf",
        "municipio",
        "segmento",
        "email_principal",
        "telefone",
        "whatsapp",
        "whatsapp_link",
        "canal_preferencial",
        "obs_contato",
    ]
    write_sheet(ws_simple, simple_headers, rows_sorted, simple_keys)

    ws1 = wb.create_sheet("03_Todos_Detalhado")
    write_sheet(ws1, headers, rows_sorted, keys)

    ws2 = wb.create_sheet("04_Com_Email")
    write_sheet(ws2, headers, with_email, keys)

    # Dedicated phone/WhatsApp sheet — prioritize those without email
    phone_rows = [r for r in rows_sorted if r.get("telefone") or r.get("whatsapp")]
    phone_rows = sorted(
        phone_rows,
        key=lambda r: (
            0 if not r["email_principal"] else 1,
            0 if r.get("whatsapp") else 1,
            r["uf"],
            r["empresa"],
        ),
    )
    wa_headers = [
        "Empresa",
        "UF",
        "Município",
        "Segmento",
        "Telefone",
        "WhatsApp",
        "Link WhatsApp",
        "E-mail (se houver)",
        "Canal Preferencial",
        "Status",
        "Observações",
    ]
    wa_keys = [
        "empresa",
        "uf",
        "municipio",
        "segmento",
        "telefone",
        "whatsapp",
        "whatsapp_link",
        "email_principal",
        "canal_preferencial",
        "status",
        "obs_contato",
    ]
    ws_wa = wb.create_sheet("05_Telefone_WhatsApp")
    write_sheet(ws_wa, wa_headers, phone_rows, wa_keys)

    ws3 = wb.create_sheet("06_Sem_Email")
    write_sheet(ws3, simple_headers, without, simple_keys)

    # Por UF
    ws4 = wb.create_sheet("07_Por_UF")
    from collections import Counter, defaultdict

    by_uf = defaultdict(lambda: {"total": 0, "com_email": 0, "sem_email": 0})
    for r in rows:
        uf = r["uf"] or "Não informado"
        by_uf[uf]["total"] += 1
        if r["email_principal"]:
            by_uf[uf]["com_email"] += 1
        else:
            by_uf[uf]["sem_email"] += 1
    uf_rows = []
    for uf, st in sorted(by_uf.items()):
        uf_rows.append(
            {
                "empresa": uf,
                "uf": st["total"],
                "municipio": st["com_email"],
                "segmento": st["sem_email"],
                "area": f"{(st['com_email']/st['total']*100):.0f}%" if st["total"] else "0%",
            }
        )
    # custom small table
    h4 = ["UF", "Total Empresas", "Com E-mail", "Sem E-mail", "% Com E-mail"]
    k4 = ["empresa", "uf", "municipio", "segmento", "area"]
    write_sheet(ws4, h4, uf_rows, k4)

    # Focus sheet
    focus_kw = (
        "ração",
        "racao",
        "nutri",
        "frigor",
        "integradora",
        "pet",
        "latic",
        "confinamento",
        "usina",
        "multinacional",
        "indústria",
        "industria",
    )
    focus = []
    for r in rows_sorted:
        blob = " ".join([r["segmento"], r["area"], r["perfil"], r["empresa"]]).lower()
        blob_n = norm(blob)
        if any(k in blob_n for k in focus_kw) or r["email_principal"]:
            # keep nutrition/frigo even without email; also keep those with email that match
            if any(k in blob_n for k in ("racao", "nutri", "frigor", "pet", "latic", "integradora", "usina", "confinamento")):
                focus.append(r)
    # dedupe preserve
    seen = set()
    focus_u = []
    for r in focus:
        if r["empresa"] in seen:
            continue
        seen.add(r["empresa"])
        focus_u.append(r)
    ws5 = wb.create_sheet("08_Nutricao_Frigorificos")
    write_sheet(ws5, simple_headers, focus_u, simple_keys)

    # Clean mailing list for outreach
    ws_mail = wb.create_sheet("09_Lista_Disparo_Email")
    mail_headers = [
        "Empresa",
        "UF",
        "Município",
        "Segmento",
        "E-mail",
        "Telefone",
        "WhatsApp",
        "Link WhatsApp",
        "Confiança",
        "Observações",
    ]
    mail_keys = [
        "empresa",
        "uf",
        "municipio",
        "segmento",
        "email_principal",
        "telefone",
        "whatsapp",
        "whatsapp_link",
        "confianca",
        "obs_contato",
    ]
    write_sheet(ws_mail, mail_headers, with_email, mail_keys)

    # Legend sheet
    ws6 = wb.create_sheet("10_Legenda")
    ws6["A1"] = "Legenda de Confiança"
    ws6["A1"].font = Font(bold=True, size=14, color="1F4E79")
    legend = [
        ("Alta", "E-mail publicado em site oficial, página de contato institucional ou documento da própria empresa."),
        ("Média", "E-mail em cadastro CNPJ, diretório setorial confiável ou canal institucional indireto (RI/imprensa) quando comercial não é público."),
        ("Baixa", "E-mail pessoal/Gmail/Hotmail publicado pela empresa, ou canal secundário (RH/vagas/recrutamento)."),
        ("-", "Sem e-mail público localizado na pesquisa."),
    ]
    ws6["A3"] = "Nível"
    ws6["B3"] = "Significado"
    for i, (a, b) in enumerate(legend, 4):
        ws6[f"A{i}"] = a
        ws6[f"B{i}"] = b
    style_header(ws6, 2)
    for col in range(1, 3):
        cell = ws6.cell(3, col)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
    ws6.column_dimensions["A"].width = 12
    ws6.column_dimensions["B"].width = 110

    wb.save(OUT_XLSX)
    # also copy to artifacts and research folder
    import shutil

    shutil.copy2(OUT_XLSX, ROOT / "ALB_Brasil_Emails_Empresas.xlsx")
    shutil.copy2(OUT_XLSX, OUT_DIR / "ALB_Brasil_Emails_Empresas.xlsx")

    # save merged json
    (ROOT / "contatos_finais.json").write_text(json.dumps(rows_sorted, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Excel: {OUT_XLSX}")
    print(f"Total: {len(rows)} | Com email: {len(with_email)} | Sem email: {len(without)}")
    print(f"Focus sheet: {len(focus_u)}")


if __name__ == "__main__":
    main()
